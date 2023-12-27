from openpyxl import load_workbook
import pandas as pd
import numpy as np

class InputData:
    def __init__(self):

        self.resource_df = None
        self.pue_input_dict = None
        self.tables_dict = None

    def dump_resource_df(self, timeseries):
        self.resource_df = self.build_resource_df(timeseries)
        self.resource_df.to_csv('./data/resource_df.csv')

    def read_resource_df_dump(self):
        self.resource_df = pd.read_csv('./data/resource_df.csv', index_col=0, parse_dates=True)

    def build_resource_df(self, timeseries):
        # Get weekly and hourly resources from input_data_dict
        weekly_resources = self.input_data_dict['weekly_resources']['df']
        weekly_resources.set_index(weekly_resources.columns[0], inplace=True)  # Make week number index
        hourly_resources = self.input_data_dict['hourly_resources']['df']

        # Create df for hourly resource data for model duration (given by timeseries)
        resource_dict = {}

        # -- Create list in dict for every hourly resource
        for index, resource_id in hourly_resources.iterrows():
            resource_dict[resource_id.iloc[0]] = np.zeros(len(timeseries))

        # -- Create list in dict for every weekly resource
        for resource_id in weekly_resources.columns:
            resource_dict[resource_id] = np.zeros(len(timeseries))

        # -- Iterate through timeseries
        i = 0  # row counter

        for timestamp in timeseries:
            hour = timestamp.hour
            day = timestamp.dayofweek
            week = timestamp.isocalendar().week

            row = []
            # -- Loop through every hourly resource in this row
            for index, resource_id in hourly_resources.iterrows():
                resource_dict[resource_id.iloc[0]][i] = self.input_data_dict[resource_id.iloc[0]]['df'].iloc[hour, day]

            # -- Loop through every weekly resource in this row
            for resource in weekly_resources.columns:
                # Set resource price for this hour based on input data (corresponding week of year)
                resource_dict[resource][i] = weekly_resources[resource].loc[week]

            i = i + 1

        resource_df = pd.DataFrame(data=resource_dict, index=timeseries)
        resource_df.reset_index(inplace=True, drop=True)  # remove datetime index

        return resource_df

    def get_all_tables(self, file):
        """ Get all tables from a given workbook. Returns a dictionary of tables.
            Requires a filename, which includes the file path and filename. """

        # Load the workbook, from the filename, setting read_only to False
        wb = load_workbook(filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

        # Initialize the dictionary of tables
        tables_dict = {}

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            # Get each table in the worksheet
            for tbl in ws.tables.values():
                # First, add some info about the table to the dictionary
                tables_dict[tbl.name] = {
                    'table_name': tbl.name,
                    'worksheet': ws_name,
                    'num_cols': len(tbl.tableColumns),
                    'table_range': tbl.ref}

                # Grab the 'data' from the table
                data = ws[tbl.ref]

                # Now convert the table 'data' to a Pandas DataFrame
                # First get a list of all rows, including the first header row
                rows_list = []
                for row in data:
                    # Get a list of all columns in each row
                    cols = []
                    for col in row:
                        cols.append(col.value)
                    rows_list.append(cols)

                # Create a pandas dataframe from the rows_list.
                # The first row is the column names
                df = pd.DataFrame(data=rows_list[1:], columns=rows_list[0])

                # If first column name is index, make index
                if df.columns[0] == 'index':
                    df = df.set_index('index', drop=True)

                # Add the dataframe to the dictionary of tables
                tables_dict[tbl.name]['df'] = df

                # If df has only one "value" column, also save as dict with {key=row_index : value}
                if len(df.columns) == 1 and df.columns[0] == 'value':
                    dct = df['value'].to_dict()
                    tables_dict[tbl.name]['dct'] = dct

        self.tables_dict = tables_dict

        return tables_dict

    def read_pue_input(self, file):
        """
        Read pue_consumer_data_input excel file

        :param file:
        :return: nested dict: entry for each appliance with dict containing appliance properties:
        {
        property_name (column in appliance_data worksheet): property_value
        ...
        weekly_preferences: df containing weekly schedule of usage times and usage preferences
        monthly_variation: df containing variation of usage_time and it's variability for each month of the year
        }

        """

        # Load the workbook, from the filename, setting read_only to False
        wb = load_workbook(filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

        # Initialize the dictionary of tables
        #pue_input_dict = {}

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:

            ws = wb[ws_name]    # select current worksheet

            if ws_name == 'general_input':   # Special treatment for general input
                print() # Currently not used
            elif ws_name == "appliance_data":    # Special treatment for appliance_data list
                # Select each table in the worksheet
                for tbl in ws.tables.values():
                    # Get table data
                    data = ws[tbl.ref]
                    # Convert table to df
                    rows_list = []
                    for row in data:
                        cols = []
                        for col in row:
                            cols.append(col.value)
                        rows_list.append(cols)

                    # Create a pandas dataframe from the rows_list.
                    # The first row is the column names
                    df = pd.DataFrame(data=rows_list[1:], columns=rows_list[0])
                    # make appliance_id index and drop empty rows
                    df = df.set_index(df['Appliance'], drop=True).dropna(how='all')
                    # Create dict entry for every appliance, each column value is dict
                    pue_input_dict = df.to_dict('index')
            else:   # for every other sheet -> get appliance input data
                # Check if this appliance was listed in appliance_data (otherwise no key exists)

                # Turn this worksheet into df
                ws = list(ws.values)
                df = pd.DataFrame(ws[1:], columns=ws[0])

                try:
                    pue_input_dict[ws_name]['weekly_preferences'] = df.iloc[:, 0:8].set_index(keys=df.columns[0], drop=True)
                    pue_input_dict[ws_name]['monthly_variation'] = df.iloc[0:12, 9:12].set_index(keys=df.columns[9], drop=True)
                except KeyError:
                    raise UserWarning(ws_name + ' is not listed in appliance_data table')

        self.pue_input_dict = pue_input_dict

        return pue_input_dict
